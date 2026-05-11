"""One-off migration: import O&M Sites.xlsx into SQLite.

Usage:
    python migrate_excel.py                      # imports from the default path
    python migrate_excel.py path/to/O&M Sites.xlsx

Re-running replaces the contents of the contracts table. Dashboard edits made
after a migration run will be LOST if you migrate again — use with care.
"""
from __future__ import annotations

import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

from config import EXCEL_SOURCE
from db import init_db, conn, insert_contract


# ----- helpers ---------------------------------------------------------------

def _to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s.startswith("="):
        # Resolve simple arithmetic formulas like "=84/12", "=37500/12", "=2668.24*1.15"
        expr = s[1:]
        try:
            return float(eval(expr, {"__builtins__": {}}, {}))
        except Exception:
            return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def _clean_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _project_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


def _status_for(row: dict, sheet: str) -> tuple[str, list[str]]:
    """Determine status and collect a list of data quality notes."""
    notes = []
    if sheet == "commercial" and row.get("invoice_type") == "No Invoice":
        return "internal_no_invoice", ["Forest Energy owned site, not billed"]

    missing = []
    if not row.get("start_date"):
        missing.append("start date")
    if row.get("base_monthly_rent") in (None, 0):
        missing.append("monthly rent")
    if row.get("contract_term_years") in (None, 0):
        missing.append("contract term")
    if sheet == "commercial" and not row.get("invoice_type"):
        missing.append("invoice type")
    if row.get("escalation_pct") is None:
        missing.append("escalation %")

    if missing:
        notes.append("Missing on import: " + ", ".join(missing))
        return "incomplete", notes
    return "active", notes


# ----- importers -------------------------------------------------------------

def _import_commercial(ws) -> list[dict]:
    out = []
    # Skip header (row 1) and any trailer totals rows (we stop at the first row with no client and no project number)
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        # Expected columns: A idx, B project, C client, D contact, E invoice type,
        # F start, G term, H monthly, I escalation
        if r is None or len(r) < 9:
            continue
        _idx, proj, client, contact, inv_type, start_d, term, rent, esc = r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8]
        if not client and not proj:
            continue  # blank separator / totals row
        # Totals footer detection: some rows only have a formula in H with no client.
        if not client:
            continue

        row = {
            "sheet_source": "commercial",
            "project_number": _project_number(proj),
            "client_name": _clean_str(client),
            "contact_person": _clean_str(contact),
            "client_email": None,
            "invoice_type": _clean_str(inv_type),
            "start_date": _to_date(start_d),
            "contract_term_years": _to_float(term),
            "base_monthly_rent": _to_float(rent),
            "escalation_pct": _to_float(esc),
            "vat_treatment": "ex_vat",
            "auto_renew": 1,
            "notice_period_days": 60,
        }
        status, notes = _status_for(row, "commercial")
        row["status"] = status
        row["notes"] = "; ".join(notes) if notes else None
        out.append(row)
    return out


def _import_residential(ws) -> list[dict]:
    out = []
    # Expected columns: A project, B client, C start, D term, E monthly, F escalation
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if r is None or len(r) < 6:
            continue
        proj, client, start_d, term, rent, esc = r[0], r[1], r[2], r[3], r[4], r[5]
        if not client and not proj:
            continue
        if not client:
            continue

        row = {
            "sheet_source": "residential",
            "project_number": _project_number(proj),
            "client_name": _clean_str(client),
            "contact_person": None,
            "client_email": None,
            "invoice_type": None,
            "start_date": _to_date(start_d),
            "contract_term_years": _to_float(term),
            "base_monthly_rent": _to_float(rent),
            "escalation_pct": _to_float(esc),
            "vat_treatment": "inc_vat",
            "auto_renew": 1,
            "notice_period_days": 60,
        }
        status, notes = _status_for(row, "residential")
        row["status"] = status
        row["notes"] = "; ".join(notes) if notes else None
        out.append(row)
    return out


def run(excel_path: str) -> dict:
    init_db()
    wb = openpyxl.load_workbook(excel_path, data_only=False)

    # Wipe existing rows so re-runs are idempotent.
    with conn() as c:
        cur = c.cursor()
        cur.execute("DELETE FROM contracts")

    imported = {"commercial": 0, "residential": 0, "incomplete": 0, "internal": 0}

    if "Forest Energy Commercial" in wb.sheetnames:
        for row in _import_commercial(wb["Forest Energy Commercial"]):
            insert_contract(row)
            imported["commercial"] += 1
            if row["status"] == "incomplete":
                imported["incomplete"] += 1
            if row["status"] == "internal_no_invoice":
                imported["internal"] += 1

    if "Forest Energy Residential PPA" in wb.sheetnames:
        for row in _import_residential(wb["Forest Energy Residential PPA"]):
            insert_contract(row)
            imported["residential"] += 1
            if row["status"] == "incomplete":
                imported["incomplete"] += 1

    return imported


if __name__ == "__main__":
    source = sys.argv[1] if len(sys.argv) > 1 else EXCEL_SOURCE
    print(f"Importing from: {source}")
    stats = run(source)
    print("Done.")
    print(f"  Commercial contracts imported:   {stats['commercial']}")
    print(f"  Residential contracts imported:  {stats['residential']}")
    print(f"  Flagged as incomplete:           {stats['incomplete']}")
    print(f"  Internal (Forest Energy owned):  {stats['internal']}")
